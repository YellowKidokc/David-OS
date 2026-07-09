"""Purpose: Extract the 20 Questions workbook into reviewable Markdown and JSON.
Date: 2026-07-08 | Author: codex | Status: TESTED via Gate 0 review pytest fixture run.

This Task 0 utility is intentionally extraction-only. It does not decide which
questions are auto-answerable and does not infer scan sources; those fields stay
null for Opus/David review after the generated bank is posted to town-square.
"""
from __future__ import annotations

import argparse
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

QUESTION_CATEGORIES = {"WHO", "WHAT", "WHEN", "WHERE", "WHY", "FRAMEWORK"}
ANSWER_TYPES = {"text", "choice", "number", "path", "bool"}
QUESTION_ID_RE = re.compile(r"\bQ(?:uestion)?\s*0*([1-9]|1\d|20)\b", re.IGNORECASE)


@dataclass(frozen=True)
class CellRecord:
    sheet: str
    coordinate: str
    row: int
    column: int
    value: str


def _stringify(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def discover_workbook(folder: Path) -> Path:
    """Return the only .xlsx workbook in a folder, or stop instead of guessing."""
    candidates = sorted(path for path in folder.glob("*.xlsx") if not path.name.startswith("~$"))
    if len(candidates) != 1:
        names = ", ".join(path.name for path in candidates) or "none"
        raise SystemExit(
            f"Expected exactly one .xlsx workbook in {folder}; found {len(candidates)} ({names}). "
            "STOP: post to town-square for filename confirmation."
        )
    return candidates[0]


def _row_values(row: Iterable[Any]) -> list[str]:
    return [_stringify(cell.value) for cell in row]


def _header_map(values: list[str]) -> dict[str, int]:
    return {value.lower().strip().replace(" ", "_"): index for index, value in enumerate(values) if value}


def _id_from_text(text: str, fallback: int) -> str:
    match = QUESTION_ID_RE.search(text)
    if match:
        return f"Q{int(match.group(1)):02d}"
    return f"Q{fallback:02d}"


def _looks_like_question(values: list[str], headers: dict[str, int]) -> bool:
    if not any(values):
        return False
    id_fields = ("id", "question_id", "qid", "q")
    if any(field in headers for field in id_fields):
        # An id column exists: it is authoritative. Rows whose id does not
        # look like a question id are misc, even if other cells have text.
        id_value = next((values[headers[field]] for field in id_fields if field in headers and headers[field] < len(values)), "")
        return bool(QUESTION_ID_RE.search(id_value) or re.fullmatch(r"0*([1-9]|1\d|20)", id_value or ""))
    question_fields = ("question", "text", "prompt")
    if any(field in headers and headers[field] < len(values) and values[headers[field]] for field in question_fields):
        return True
    joined = " | ".join(values)
    return bool(QUESTION_ID_RE.search(joined) and "?" in joined)


def _field(values: list[str], headers: dict[str, int], names: tuple[str, ...]) -> str:
    for name in names:
        index = headers.get(name)
        if index is not None and index < len(values):
            return values[index]
    return ""


def _normalize_category(raw: str, text: str) -> str:
    value = raw.upper().strip()
    if value in QUESTION_CATEGORIES:
        return value
    upper_text = text.upper()
    for category in QUESTION_CATEGORIES:
        if upper_text.startswith(category) or f"[{category}]" in upper_text:
            return category
    return "WHAT"


def _normalize_answer_type(raw: str) -> str:
    value = raw.lower().strip()
    if value in ANSWER_TYPES:
        return value
    if "yes" in value or "bool" in value:
        return "bool"
    if "path" in value or "folder" in value:
        return "path"
    if "number" in value or "count" in value:
        return "number"
    if "choice" in value or "select" in value:
        return "choice"
    return "text"


def _choices(raw: str) -> list[str]:
    if not raw:
        return []
    return [part.strip() for part in re.split(r"\s*(?:\||,|;)\s*", raw) if part.strip()]


def extract_workbook(workbook_path: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[CellRecord], dict[str, int]]:
    wb = load_workbook(workbook_path, data_only=False, read_only=True)
    questions: list[dict[str, Any]] = []
    misc_rows: list[dict[str, Any]] = []
    cells: list[CellRecord] = []
    question_counter = 0
    stats = {"sheets": len(wb.sheetnames), "rows": 0, "non_empty_cells": 0, "questions": 0, "misc_items": 0}

    for ws in wb.worksheets:
        rows = list(ws.iter_rows())
        header_values = _row_values(rows[0]) if rows else []
        headers = _header_map(header_values)
        for row_idx, row in enumerate(rows, start=1):
            values = _row_values(row)
            if not any(values):
                continue
            stats["rows"] += 1
            row_cells = [CellRecord(ws.title, getattr(cell, "coordinate", f"R{row_idx}C{col_idx}"), row_idx, col_idx, _stringify(cell.value)) for col_idx, cell in enumerate(row, start=1) if _stringify(cell.value)]
            cells.extend(row_cells)
            stats["non_empty_cells"] += len(row_cells)
            if row_idx == 1 and headers:
                misc_rows.append({"sheet": ws.title, "row": row_idx, "cells": [cell.__dict__ for cell in row_cells], "reason": "header"})
                continue
            if _looks_like_question(values, headers):
                question_counter += 1
                text = _field(values, headers, ("question", "text", "prompt")) or next((value for value in values if "?" in value), " | ".join(v for v in values if v))
                raw_id = _field(values, headers, ("id", "question_id", "qid", "q"))
                qid = _id_from_text(raw_id or text, question_counter)
                notes_parts = []
                for index, value in enumerate(values):
                    if not value or value in {raw_id, text}:
                        continue
                    header = header_values[index] if index < len(header_values) and header_values[index] else f"Column {index + 1}"
                    if header.lower().strip().replace(" ", "_") in {"category", "answer_type", "choices"}:
                        continue
                    notes_parts.append(f"{header}: {value}")
                questions.append({
                    "id": qid,
                    "text": text,
                    "category": _normalize_category(_field(values, headers, ("category", "bucket", "type")), text),
                    "answer_type": _normalize_answer_type(_field(values, headers, ("answer_type", "answer", "response_type"))),
                    "choices": _choices(_field(values, headers, ("choices", "options", "choice"))),
                    "auto_answerable": None,
                    "scan_source": None,
                    "notes": "\n".join(notes_parts),
                    "source": {"sheet": ws.title, "row": row_idx, "cells": [cell.__dict__ for cell in row_cells]},
                })
            else:
                misc_rows.append({"sheet": ws.title, "row": row_idx, "cells": [cell.__dict__ for cell in row_cells], "reason": "not clearly a question"})
    stats["questions"] = len(questions)
    stats["misc_items"] = len(misc_rows)
    return questions, misc_rows, cells, stats


def write_outputs(workbook_path: Path, output_dir: Path) -> dict[str, Path]:
    questions, misc_rows, cells, stats = extract_workbook(workbook_path)
    output_dir.mkdir(parents=True, exist_ok=True)
    bank_md = output_dir / "20Q_QUESTION_BANK.md"
    bank_json = output_dir / "20q_bank.json"
    misc_md = output_dir / "20Q_WORKBOOK_MISC.md"

    bank_md.write_text(_render_bank_md(workbook_path, questions, stats), encoding="utf-8")
    bank_json.write_text(json.dumps({"schema": "david-os.q20_bank.v1", "source_workbook": str(workbook_path), "stats": stats, "questions": questions}, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    misc_md.write_text(_render_misc_md(workbook_path, misc_rows, stats), encoding="utf-8")
    _validate_round_trip(cells, questions, misc_rows)
    return {"bank_md": bank_md, "bank_json": bank_json, "misc_md": misc_md}


def _render_bank_md(workbook_path: Path, questions: list[dict[str, Any]], stats: dict[str, int]) -> str:
    lines = ["# 20Q Question Bank", "", "Purpose: Human-readable extraction from the 20 Questions workbook.", "Date: 2026-07-08 | Generated by: codex | Status: TESTED", f"Source workbook: `{workbook_path}`", "", f"Extraction summary: {stats['sheets']} sheets, {stats['rows']} non-empty rows, {stats['questions']} questions extracted, {stats['misc_items']} misc items.", ""]
    for q in questions:
        lines += [f"## {q['id']}", "", f"- Text: {q['text']}", f"- Category: {q['category']}", f"- Answer type: {q['answer_type']}", f"- Choices: {', '.join(q['choices']) if q['choices'] else '[]'}", "- Auto-answerable: null", "- Scan source: null", f"- Source: {q['source']['sheet']} row {q['source']['row']}"]
        if q["notes"]:
            lines += ["", "### Notes", q["notes"]]
        lines.append("")
    return "\n".join(lines)


def _render_misc_md(workbook_path: Path, misc_rows: list[dict[str, Any]], stats: dict[str, int]) -> str:
    lines = ["# 20Q Workbook Miscellaneous Material", "", "Purpose: Verbatim non-question workbook material preserved for review.", "Date: 2026-07-08 | Generated by: codex | Status: TESTED", f"Source workbook: `{workbook_path}`", "", f"Extraction summary: {stats['sheets']} sheets, {stats['rows']} non-empty rows, {stats['questions']} questions extracted, {stats['misc_items']} misc items.", ""]
    for item in misc_rows:
        lines += [f"## {item['sheet']} row {item['row']}", "", f"Reason: {item['reason']}", "", "| Cell | Value |", "| --- | --- |"]
        for cell in item["cells"]:
            value = str(cell["value"]).replace("|", "\\|").replace("\n", "<br>")
            lines.append(f"| {cell['coordinate']} | {value} |")
        lines.append("")
    return "\n".join(lines)


def _validate_round_trip(cells: list[CellRecord], questions: list[dict[str, Any]], misc_rows: list[dict[str, Any]]) -> None:
    covered = {(cell["sheet"], cell["coordinate"]) for q in questions for cell in q["source"]["cells"]}
    covered.update((cell["sheet"], cell["coordinate"]) for item in misc_rows for cell in item["cells"])
    original = {(cell.sheet, cell.coordinate) for cell in cells}
    missing = original - covered
    if missing:
        raise RuntimeError(f"Round-trip validation failed; missing cells: {sorted(missing)[:10]}")


def main() -> int:
    parser = argparse.ArgumentParser(description="Extract the 20 Questions workbook for Task 0 review.")
    parser.add_argument("workbook_or_folder", help="Path to the .xlsx workbook, or a folder containing exactly one .xlsx file.")
    parser.add_argument("--out", default="core", help="Output directory for 20Q_QUESTION_BANK.md, 20q_bank.json, and 20Q_WORKBOOK_MISC.md.")
    args = parser.parse_args()
    source = Path(args.workbook_or_folder)
    workbook = discover_workbook(source) if source.is_dir() else source
    outputs = write_outputs(workbook, Path(args.out))
    data = json.loads(outputs["bank_json"].read_text(encoding="utf-8"))
    stats = data["stats"]
    print(
        "Task 0 extraction complete: "
        f"{stats['sheets']} sheets, {stats['rows']} non-empty rows, "
        f"{stats['questions']} questions extracted, {stats['misc_items']} misc items, "
        f"{stats['non_empty_cells']} non-empty cells covered."
    )
    print("POST TO TOWN-SQUARE AND WAIT FOR REVIEW BEFORE TASK 1.")
    for label, path in outputs.items():
        print(f"{label}: {path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
