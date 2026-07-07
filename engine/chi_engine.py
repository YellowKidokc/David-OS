#!/usr/bin/env python3
"""
χ / Qi v5 Statistical Audit Engine — scaffold
=============================================

This is the v5 layer that sits between:
- filetagger_chi_v2.py: fast corpus scanner
- chi_jm_diagnostic_engine_v4_gold.py: deep paper/claim audit
- .chi sidecars: readable audit records
- SQLite: statistical memory
- HTML dashboards: corpus/page/sentence navigation

Core rule:
Every label must have a score.
Every score must have evidence spans.
Every evidence span must have a weight.
Every winner must beat a rival.
Every weakness must produce a symptom.
Every symptom must produce a repair path.

This file is intentionally standard-library first. If openpyxl is present, it can
load David's lexicon workbook. If not, it falls back to built-in seed terms.
"""
from __future__ import annotations

import argparse
import dataclasses
import hashlib
import html
import json
import math
import os
import re
import sqlite3
import statistics
from collections import Counter, defaultdict
from dataclasses import dataclass, field, asdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

ENGINE_VERSION = "chi_qi_v5.0-scaffold"

TOKEN_RE = re.compile(r"[A-Za-z][A-Za-z0-9_\-']+|[χΧΩΣΘΛΨΦ][\w_]*")
PARA_RE = re.compile(r"\n\s*\n+")
SENT_RE = re.compile(r"(?<=[.!?])\s+(?=[A-Z0-9\"'“‘])")
HTML_TAG_RE = re.compile(r"(?s)<[^>]+>")
SCRIPT_STYLE_RE = re.compile(r"(?is)<(script|style).*?>.*?</\1>")

# Seed terms are deliberately small. The lexicon workbook should be primary.
SEED_COLLECTIONS: Dict[str, List[str]] = {
    "EVIDENCE_TERMS": ["source", "citation", "data", "measurement", "study", "observed", "figure", "table", "reproducible", "peer-reviewed"],
    "BOUNDARY_TERMS": ["does not prove", "not claim", "limited", "scope", "boundary", "candidate", "hypothesis", "analogy", "metaphor", "within the model"],
    "FALSIFY_TERMS": ["kill condition", "falsify", "would fail", "counterexample", "refuted", "death condition", "wrong if"],
    "COHERENCE_TERMS": ["coherence", "integrate", "alignment", "structure", "mechanism", "invariant", "logos", "unify"],
    "DISCOHERENCE_TERMS": ["contradiction", "incoherent", "fragmentation", "drift", "unsupported", "circular", "ad hoc"],
    "FRUITS": ["love", "joy", "peace", "patience", "kindness", "goodness", "faithfulness", "gentleness", "self-control"],
    "ANTI_FRUITS": ["rage", "domination", "coercion", "contempt", "dehumanization", "panic", "manipulation", "despair"],
    "LAW_KEYWORDS": ["gravity", "grace", "motion", "will", "light", "truth", "strong force", "love", "entropy", "justice", "information", "logos", "quantum", "faith", "relativity", "relationship", "weak force", "decay", "coherence", "christ"],
}

MASTER_VARIABLES = {
    "G": ["grace", "mercy", "restore", "redemption", "forgive"],
    "M": ["will", "repentance", "moral", "alignment", "choice"],
    "E": ["truth", "evidence", "source", "measure", "citation"],
    "S": ["entropy", "judgment", "decay", "consequence", "collapse"],
    "T": ["time", "kairos", "sequence", "history", "threshold"],
    "K": ["knowledge", "information", "logos", "signal", "model"],
    "R": ["relationship", "family", "covenant", "binding", "community"],
    "Q": ["quantum", "faith", "observer", "measurement", "uncertainty"],
    "F": ["force", "action", "motion", "will", "moral force"],
    "C": ["coherence", "christ", "integration", "shalom", "kingdom"],
}

DOMAIN_TERMS = {
    "theology": ["god", "christ", "jesus", "spirit", "grace", "sin", "salvation", "scripture", "logos"],
    "physics": ["physics", "quantum", "field", "entropy", "relativity", "gravity", "mass", "energy", "measurement"],
    "information_theory": ["information", "signal", "noise", "channel", "shannon", "compression", "encoding"],
    "formal_math": ["equation", "theorem", "proof", "lean", "operator", "function", "variable", "axiom"],
    "history": ["historical", "century", "document", "record", "timeline", "1900", "1919"],
    "epistemology": ["truth", "knowledge", "ground", "regress", "assumption", "falsifiability", "method"],
    "ethics_moral_philosophy": ["good", "evil", "justice", "ought", "virtue", "responsibility", "accountability"],
    "sociology_culture": ["society", "institution", "culture", "community", "family", "civilization"],
    "psychology": ["behavior", "emotion", "anxiety", "habit", "trauma", "identity", "willpower"],
    "ai_methodology": ["ai", "llm", "model", "prompt", "agent", "classification", "corpus"],
}

@dataclass
class SemanticTerm:
    term: str
    collection: str
    bucket: str = ""
    subbucket: str = ""
    polarity: str = "positive"
    weight: float = 1.0
    danger_level: str = "low"
    source_sheet: str = ""

@dataclass
class TextUnit:
    unit_id: str
    file_id: str
    unit_type: str
    ordinal: int
    text: str
    anchor: str
    start_char: int
    end_char: int
    parent_unit_id: Optional[str] = None
    heading: str = ""

@dataclass
class EvidenceSpan:
    span_id: str
    result_id: str
    file_id: str
    unit_id: str
    metric_id: str
    term: str
    normalized_term: str
    lexicon_collection: str
    lexicon_bucket: str
    lexicon_subbucket: str
    polarity: str
    danger_level: str
    term_weight: float
    contribution: float
    matched_text: str
    paragraph_index: int
    sentence_index: int
    start_char: int
    end_char: int
    context_before: str = ""
    context_after: str = ""

@dataclass
class MetricResult:
    result_id: str
    file_id: str
    unit_id: str
    scope: str
    metric_id: str
    metric_family: str
    metric_object: str
    facet: str
    raw_hits: int = 0
    weighted_hits: float = 0.0
    unique_terms: int = 0
    density_per_1000_words: float = 0.0
    distribution_score: float = 0.0
    cooccurrence_score: float = 0.0
    counter_signal_score: float = 0.0
    structural_completeness: float = 0.0
    evidence_support: float = 0.0
    boundary_support: float = 0.0
    formal_support: float = 0.0
    kill_condition_support: float = 0.0
    repair_support: float = 0.0
    rival_metric_id: Optional[str] = None
    rival_score: float = 0.0
    margin: float = 0.0
    score: float = 0.0
    confidence: float = 0.0
    normalized_pct: float = 0.0
    qualifier: str = "absent"
    symptom: str = ""
    symptom_severity: str = "none"
    repair_action: str = ""
    route_trigger: bool = False
    evidence_spans: List[EvidenceSpan] = field(default_factory=list)

def clamp(x: float, lo: float = 0.0, hi: float = 100.0) -> float:
    try:
        if math.isnan(x):
            return lo
    except Exception:
        return lo
    return max(lo, min(hi, float(x)))

def pct(x: float) -> float:
    return round(clamp(x), 2)

def sha256_text(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8", errors="ignore")).hexdigest()

def tokens(text: str) -> List[str]:
    return [t.lower() for t in TOKEN_RE.findall(text or "")]

def normalize_text(text: str) -> str:
    return re.sub(r"\s+", " ", text or "").strip()

def read_text(path: Path) -> str:
    raw = path.read_text(encoding="utf-8", errors="replace")
    if path.suffix.lower() in {".html", ".htm"}:
        raw = SCRIPT_STYLE_RE.sub(" ", raw)
        raw = re.sub(r"(?is)</(p|div|section|article|li|h[1-6]|tr)>", "\n", raw)
        raw = HTML_TAG_RE.sub(" ", raw)
        raw = html.unescape(raw)
    return raw.replace("\r\n", "\n").replace("\r", "\n")

def split_paragraphs_with_offsets(text: str, file_id: str) -> List[TextUnit]:
    units: List[TextUnit] = []
    ordinal = 0
    for m in re.finditer(r"\S(?:.*?)(?=\n\s*\n+|\Z)", text, flags=re.DOTALL):
        para = normalize_text(m.group(0))
        if len(para.split()) < 4:
            continue
        ordinal += 1
        units.append(TextUnit(
            unit_id=f"{file_id}::p{ordinal:04d}",
            file_id=file_id,
            unit_type="paragraph",
            ordinal=ordinal,
            text=para,
            anchor=f"p-{ordinal:04d}",
            start_char=m.start(),
            end_char=m.end(),
        ))
    return units

def split_sentences_from_paragraphs(paras: List[TextUnit]) -> List[TextUnit]:
    out: List[TextUnit] = []
    ordinal = 0
    for p in paras:
        offset_cursor = p.start_char
        for piece in SENT_RE.split(p.text):
            sent = normalize_text(piece)
            if len(sent) < 20:
                continue
            ordinal += 1
            # approximate local offset; enough for the first v5 layer
            idx = p.text.find(sent)
            start = p.start_char + max(0, idx)
            out.append(TextUnit(
                unit_id=f"{p.file_id}::s{ordinal:04d}",
                file_id=p.file_id,
                unit_type="sentence",
                ordinal=ordinal,
                text=sent,
                anchor=f"s-{ordinal:04d}",
                start_char=start,
                end_char=start + len(sent),
                parent_unit_id=p.unit_id,
            ))
    return out

class LexiconStore:
    def __init__(self, path: Optional[Path] = None, policy: str = "merge"):
        self.path = str(path) if path else ""
        self.policy = policy
        self.loaded = False
        self.warnings: List[str] = []
        self.terms_by_collection: Dict[str, List[SemanticTerm]] = defaultdict(list)
        self.sheet_names: List[str] = []
        self.checksum: str = ""

    @classmethod
    def load(cls, path: Optional[Path], policy: str = "merge") -> "LexiconStore":
        store = cls(path, policy)
        for collection, terms in SEED_COLLECTIONS.items():
            for t in terms:
                store.terms_by_collection[collection].append(SemanticTerm(term=t, collection=collection, source_sheet="built_in"))
        if not path:
            return store
        try:
            store.checksum = sha256_text(path.read_bytes().hex())
            from openpyxl import load_workbook  # optional runtime dependency
            wb = load_workbook(path, read_only=True, data_only=True)
            store.sheet_names = list(wb.sheetnames)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = ws.iter_rows(values_only=True)
                try:
                    headers = [str(h).strip() if h is not None else "" for h in next(rows)]
                except StopIteration:
                    continue
                h = {name: i for i, name in enumerate(headers) if name}
                def cell(row, key, default=""):
                    idx = h.get(key)
                    if idx is None or idx >= len(row):
                        return default
                    return row[idx] if row[idx] is not None else default
                for row in rows:
                    if row is None or not any(str(x).strip() for x in row if x is not None):
                        continue
                    collection = str(cell(row, "collection", sheet_name) or sheet_name).strip()
                    term = str(cell(row, "term", "") or "").strip()
                    if not term and "key" in h and "value" in h:
                        term = str(cell(row, "value", "") or cell(row, "key", "") or "").strip()
                    if not term:
                        continue
                    try:
                        weight = float(cell(row, "weight", 1) or 1)
                    except Exception:
                        weight = 1.0
                    store.terms_by_collection[collection].append(SemanticTerm(
                        term=term,
                        collection=collection,
                        bucket=str(cell(row, "bucket", collection) or collection),
                        subbucket=str(cell(row, "subbucket", "") or ""),
                        polarity=str(cell(row, "polarity", "positive") or "positive"),
                        weight=weight,
                        danger_level=str(cell(row, "danger_level", "low") or "low"),
                        source_sheet=sheet_name,
                    ))
            store.loaded = True
        except Exception as exc:
            store.warnings.append(f"lexicon load failed: {exc}")
        return store

    def terms(self, collections: Iterable[str]) -> List[SemanticTerm]:
        out: List[SemanticTerm] = []
        for c in collections:
            out.extend(self.terms_by_collection.get(c, []))
        return out

def danger_modifier(level: str) -> float:
    return {"low": 1.0, "medium": 1.15, "high": 1.35, "critical": 1.65}.get((level or "").lower(), 1.0)

def polarity_modifier(polarity: str) -> float:
    return -1.0 if (polarity or "").lower() in {"negative", "counter"} else 1.0

def term_matches(text: str, term: str) -> List[Tuple[int, int, str]]:
    if not term:
        return []
    pattern = re.escape(term.lower())
    low = text.lower()
    matches = []
    if " " in term:
        start = 0
        while True:
            i = low.find(term.lower(), start)
            if i < 0:
                break
            matches.append((i, i + len(term), text[i:i+len(term)]))
            start = i + len(term)
    else:
        for m in re.finditer(rf"\b{pattern}\b", low):
            matches.append((m.start(), m.end(), text[m.start():m.end()]))
    return matches

def make_result_id(file_id: str, unit_id: str, metric_id: str) -> str:
    return hashlib.sha1(f"{file_id}|{unit_id}|{metric_id}".encode()).hexdigest()[:20]

def build_metric_from_terms(
    file_id: str,
    unit: TextUnit,
    metric_id: str,
    family: str,
    obj: str,
    facet: str,
    terms: List[SemanticTerm],
    support_terms: List[SemanticTerm],
    counter_terms: List[SemanticTerm],
) -> MetricResult:
    result_id = make_result_id(file_id, unit.unit_id, metric_id)
    word_count = max(1, len(tokens(unit.text)))
    spans: List[EvidenceSpan] = []
    unique = set()
    weighted = 0.0
    for st in terms:
        for start, end, matched in term_matches(unit.text, st.term):
            contrib = max(0.25, st.weight) * danger_modifier(st.danger_level) * abs(polarity_modifier(st.polarity))
            weighted += contrib
            unique.add(st.term.lower())
            spans.append(EvidenceSpan(
                span_id=hashlib.sha1(f"{result_id}|{st.term}|{start}|{end}".encode()).hexdigest()[:20],
                result_id=result_id,
                file_id=file_id,
                unit_id=unit.unit_id,
                metric_id=metric_id,
                term=st.term,
                normalized_term=st.term.lower(),
                lexicon_collection=st.collection,
                lexicon_bucket=st.bucket or st.collection,
                lexicon_subbucket=st.subbucket,
                polarity=st.polarity,
                danger_level=st.danger_level,
                term_weight=st.weight,
                contribution=round(contrib, 4),
                matched_text=matched,
                paragraph_index=unit.ordinal if unit.unit_type == "paragraph" else 0,
                sentence_index=unit.ordinal if unit.unit_type == "sentence" else 0,
                start_char=unit.start_char + start,
                end_char=unit.start_char + end,
            ))
    support_score = min(100.0, len({s.term.lower() for s in support_terms if term_matches(unit.text, s.term)}) * 14)
    counter_score = min(100.0, len({s.term.lower() for s in counter_terms if term_matches(unit.text, s.term)}) * 16)
    density = 1000.0 * weighted / word_count
    marker_strength = min(100.0, weighted * 12)
    distribution = 100.0 if spans else 0.0
    score = pct(0.35 * marker_strength + 0.20 * min(100, density * 10) + 0.20 * support_score + 0.15 * distribution - 0.20 * counter_score)
    confidence = pct(0.55 * score + 0.25 * min(100, len(unique) * 20) + 0.20 * (100 - counter_score))
    qualifier = "dominant" if score >= 85 else "strong" if score >= 65 else "moderate" if score >= 35 else "weak" if score >= 10 else "absent"
    symptom = "No meaningful signal detected." if score < 10 else f"{obj} signal is {qualifier} at {unit.unit_type} level."
    if counter_score > 40:
        symptom += " Counter-signal pressure is high; inspect contradiction or mixed category."
    repair = "No immediate repair required; preserve and compare to rivals."
    if score < 35:
        repair = "Add clearer terms, evidence spans, or downgrade this label."
    if counter_score > 40:
        repair = "Split claim or add boundary language before promoting."
    return MetricResult(
        result_id=result_id,
        file_id=file_id,
        unit_id=unit.unit_id,
        scope=unit.unit_type,
        metric_id=metric_id,
        metric_family=family,
        metric_object=obj,
        facet=facet,
        raw_hits=len(spans),
        weighted_hits=round(weighted, 4),
        unique_terms=len(unique),
        density_per_1000_words=round(density, 4),
        distribution_score=distribution,
        cooccurrence_score=support_score,
        counter_signal_score=counter_score,
        evidence_support=support_score if family in {"evidence", "domain", "law"} else 0,
        boundary_support=support_score if "boundary" in metric_id else 0,
        score=score,
        confidence=confidence,
        qualifier=qualifier,
        symptom=symptom,
        symptom_severity="high" if counter_score > 60 else "medium" if score < 35 and len(spans) else "low",
        repair_action=repair,
        route_trigger=bool(score >= 70 or counter_score >= 50 or (score >= 45 and confidence < 55)),
        evidence_spans=spans,
    )

def normalize_vector(raw: Dict[str, float]) -> Dict[str, float]:
    positives = {k: max(0.0, float(v)) for k, v in raw.items() if float(v) > 0}
    total = sum(positives.values())
    if total <= 0:
        return {k: 0.0 for k in raw}
    return {k: round(v / total * 100.0, 2) for k, v in positives.items()}

def extract_profiles(file_id: str, text: str, lex: LexiconStore) -> Dict[str, Dict[str, float]]:
    full_unit = TextUnit(file_id=file_id, unit_id=file_id, unit_type="file", ordinal=0, text=text, anchor="", start_char=0, end_char=len(text))
    raw_domains = {}
    for d, terms in DOMAIN_TERMS.items():
        sts = [SemanticTerm(term=t, collection="DOMAIN_KEYWORDS", bucket=d) for t in terms]
        res = build_metric_from_terms(file_id, full_unit, f"domain.{d}.score", "domain", d, "score", sts, [], [])
        raw_domains[d] = res.score
    raw_chi = {}
    for var, terms in MASTER_VARIABLES.items():
        sts = [SemanticTerm(term=t, collection="ME_VARS", bucket=var) for t in terms]
        res = build_metric_from_terms(file_id, full_unit, f"me.{var.lower()}.score", "master_equation", var, "score", sts, [], [])
        raw_chi[var] = res.score
    return {
        "domain_vector": normalize_vector(raw_domains),
        "chi_vector": normalize_vector(raw_chi),
    }

def write_chi_sidecar(path: Path, report: Dict[str, Any], out_path: Path) -> None:
    text = ["---", "template_version: chi_qi_v5.0", f"file_id: {report['identity']['file_id']}", f"source_path: {json.dumps(str(path))}", "---", ""]
    text.append("# χ / Qi v5 File Intelligence Sidecar\n")
    text.append("## Identity\n")
    text.append("```json\n" + json.dumps(report["identity"], indent=2) + "\n```\n")
    text.append("## Lexicon\n")
    text.append("```json\n" + json.dumps(report["lexicon"], indent=2) + "\n```\n")
    text.append("## Profiles\n")
    text.append("```json\n" + json.dumps(report["profiles"], indent=2) + "\n```\n")
    text.append("## Top Metric Results\n")
    text.append("```json\n" + json.dumps(report["metric_results"][:25], indent=2) + "\n```\n")
    text.append("## Routing\n")
    text.append("```json\n" + json.dumps(report["routing"], indent=2) + "\n```\n")
    out_path.write_text("\n".join(text), encoding="utf-8")

def init_db(db_path: Path, schema_path: Optional[Path] = None) -> sqlite3.Connection:
    con = sqlite3.connect(db_path)
    if schema_path and schema_path.exists():
        con.executescript(schema_path.read_text(encoding="utf-8"))
    return con

def process_file(path: Path, lex: LexiconStore, out_dir: Path) -> Dict[str, Any]:
    raw = read_text(path)
    file_id = sha256_text(str(path) + raw[:1000])[:16]
    paras = split_paragraphs_with_offsets(raw, file_id)
    sents = split_sentences_from_paragraphs(paras)
    profiles = extract_profiles(file_id, raw, lex)

    metric_results: List[MetricResult] = []
    # Sentence-level evidence / boundary / falsification / coherence examples.
    metric_specs = [
        ("evidence.general.score", "evidence", "evidence", "score", ["EVIDENCE_TERMS"], ["BOUNDARY_TERMS"], []),
        ("boundary.general.score", "boundary", "boundary", "score", ["BOUNDARY_TERMS"], ["EVIDENCE_TERMS"], []),
        ("falsification.general.score", "falsification", "falsification", "score", ["FALSIFY_TERMS"], ["EVIDENCE_TERMS"], []),
        ("coherence.general.score", "coherence", "coherence", "score", ["COHERENCE_TERMS"], ["EVIDENCE_TERMS", "FALSIFY_TERMS"], ["DISCOHERENCE_TERMS"]),
        ("fruit.general.score", "fruit", "fruit", "score", ["FRUITS"], ["BOUNDARY_TERMS"], ["ANTI_FRUITS"]),
        ("anti_fruit.general.score", "anti_fruit", "anti_fruit", "score", ["ANTI_FRUITS"], [], ["FRUITS"]),
    ]
    units = sents[:300] if sents else paras[:300]
    for unit in units:
        for metric_id, fam, obj, facet, cols, support_cols, counter_cols in metric_specs:
            terms = lex.terms(cols)
            support = lex.terms(support_cols)
            counter = lex.terms(counter_cols)
            result = build_metric_from_terms(file_id, unit, metric_id, fam, obj, facet, terms, support, counter)
            if result.raw_hits or result.score >= 10:
                metric_results.append(result)

    high_risk = [r for r in metric_results if r.route_trigger or r.counter_signal_score > 50]
    identity = {
        "file_id": file_id,
        "source_path": str(path),
        "title": path.stem,
        "content_type": "html" if path.suffix.lower() in {".html", ".htm"} else "text",
        "extension": path.suffix.lower(),
        "word_count": len(tokens(raw)),
        "paragraph_count": len(paras),
        "sentence_count": len(sents),
        "hash_sha256": sha256_text(raw),
    }
    lexicon_block = {
        "loaded": lex.loaded,
        "path": lex.path,
        "policy": lex.policy,
        "checksum": lex.checksum,
        "sheets_loaded": lex.sheet_names,
        "role": "canonical_measurement_language",
        "warnings": lex.warnings,
    }
    report = {
        "template_version": "chi_qi_v5.0",
        "identity": identity,
        "lexicon": lexicon_block,
        "profiles": profiles,
        "metric_results": [
            {k: v for k, v in asdict(r).items() if k != "evidence_spans"} | {
                "evidence_spans": [asdict(s) for s in r.evidence_spans[:5]]
            }
            for r in sorted(metric_results, key=lambda x: (x.score, x.confidence), reverse=True)[:100]
        ],
        "routing": {
            "pass_1_complete": True,
            "pass_2_deep_audit_recommended": bool(high_risk or max((r.score for r in metric_results), default=0) >= 70),
            "pass_3_llm_review_recommended": False,
            "reasons": sorted(set([r.symptom for r in high_risk[:10]])),
        },
    }
    out_dir.mkdir(parents=True, exist_ok=True)
    write_chi_sidecar(path, report, out_dir / f"{path.name}.chi")
    (out_dir / f"{path.stem}.chi_v5.json").write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
    return report

def main(argv: Optional[List[str]] = None) -> int:
    p = argparse.ArgumentParser(description="χ / Qi v5 Statistical Audit scaffold")
    p.add_argument("--input", required=True, help="Input file or folder")
    p.add_argument("--out", required=True, help="Output folder")
    p.add_argument("--lexicon", help="Optional lexicons_master_enhanced.xlsx")
    p.add_argument("--recursive", action="store_true")
    p.add_argument("--max-files", type=int, default=0)
    args = p.parse_args(argv)

    input_path = Path(args.input).expanduser().resolve()
    out_dir = Path(args.out).expanduser().resolve()
    lex = LexiconStore.load(Path(args.lexicon).expanduser().resolve() if args.lexicon else None)

    if input_path.is_file():
        files = [input_path]
    else:
        globber = input_path.rglob("*") if args.recursive else input_path.glob("*")
        files = [x for x in globber if x.suffix.lower() in {".md", ".txt", ".html", ".htm"}]
    if args.max_files:
        files = files[:args.max_files]

    reports = []
    for i, f in enumerate(files, 1):
        print(f"[{i}/{len(files)}] {f}")
        reports.append(process_file(f, lex, out_dir))

    summary = {
        "engine_version": ENGINE_VERSION,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "file_count": len(reports),
        "lexicon_loaded": lex.loaded,
        "reports": [{"file_id": r["identity"]["file_id"], "path": r["identity"]["source_path"], "profiles": r["profiles"], "routing": r["routing"]} for r in reports],
    }
    (out_dir / "chi_qi_v5_batch_summary.json").write_text(json.dumps(summary, indent=2), encoding="utf-8")
    print(f"Done: {out_dir}")
    return 0

if __name__ == "__main__":
    raise SystemExit(main())
